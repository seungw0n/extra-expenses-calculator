# @description: Collects all functions related to handling files
# @author: seungw0n
# @installed-packages: openpyxl

from openpyxl import load_workbook, Workbook, worksheet
from datetime import datetime


def openExcel(filename: str) -> list:
    """ filename must contain extension """
    wb = load_workbook(filename=filename)
    return wb, wb.sheetnames


def readNeis(wb: Workbook) -> dict:
    sheet = wb.active

    result = dict()

    values = list(sheet.values)

    for i in range(2, len(values)):
        date = values[i][3]
        value = [values[i][2], values[i][8], values[i][9]]  # 이름, 시작시간, 종료시간

        if date in result:
            result[date].append(value)
        else:
            result[date] = [value]

    return result


def readLedger(wb: Workbook, sheetnames: list) -> dict:
    result = dict()

    for sheetname in sheetnames:
        sheet = wb[sheetname]
        values = list(sheet.values)

        for i in range(1, len(values)):
            date = values[i][0]
            price = values[i][1]
            names = values[i][2]

            if date is None:
                break

            if date in result:
                result[date].append([sheetname, names, price])
            else:
                result[date] = [[sheetname, names, price]]
    return result


# wb, sheetnames = openExcel("../files/202209/9월 특근매식비장부.xlsx")
# result = readLedger(wb, sheetnames)
# cnt = 0
# for k, v in result.items():
#     print("날짜: " + k)
#     print("\t", v)
#     cnt += len(v)
#
# print(cnt)