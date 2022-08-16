# @description: Collects all functions related to handling files
# @author: seungw0n
# @installed-packages: openpyxl

from openpyxl import load_workbook, Workbook

def read(location: str, filename: str):
    """ filename must contain extension """
    wb = load_workbook(filename = location + filename)
    sheets = wb.sheetnames

    print(sheets)


if __name__ == '__main__':
    read("./", "7월 특근매식비.xlsx")
